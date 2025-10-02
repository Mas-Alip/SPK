from openpyxl import Workbook
import sqlite3
import os

wb = Workbook()

# --- load criteria names from app database (fallback to sensible defaults) ---
db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'spk_kredit.db')
crit_names = []
try:
    if os.path.exists(db_path):
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("SELECT nama FROM kriteria ORDER BY id")
        rows = cur.fetchall()
        conn.close()
        crit_names = [r[0] for r in rows if r and r[0]]
except Exception:
    crit_names = []

# fallback if DB empty or not present
if not crit_names:
    crit_names = ["Usia", "Pendapatan", "Pekerjaan", "Jaminan"]

n = len(crit_names)
ws = wb.active
ws.title = "Criteria_Pairwise"

# headers
for j, name in enumerate(crit_names, start=2):
    ws.cell(row=1, column=j, value=name)
    ws.cell(row=j, column=1, value=name)

# sample pairwise (editable) - default identity
for i in range(n):
    for j in range(n):
        r = 2 + i
        c = 2 + j
        if i == j:
            ws.cell(row=r, column=c, value=1)
        else:
            # leave sample 1 for user to edit
            ws.cell(row=r, column=c, value=1)

# column sums at row n+3
col_sum_row = 2 + n + 1
for j in range(n):
    col_letter = None
    # place formula SUM(B2:B{n+1})
    col = 2 + j
    ws.cell(row=col_sum_row, column=col, value=f"=SUM({ws.cell(row=2, column=col).coordinate}:{ws.cell(row=1+n, column=col).coordinate})")
ws.cell(row=col_sum_row, column=1, value="Column Sum")

# --- Criteria Normalized sheet ---
cn = wb.create_sheet("Criteria_Normalized")
# headers
for j, name in enumerate(crit_names, start=2):
    cn.cell(row=1, column=j, value=name)
    cn.cell(row=j, column=1, value=name)

# normalized = Criteria_Pairwise!cell / Criteria_Pairwise!col_sum_row
for i in range(n):
    for j in range(n):
        r = 2 + i
        c = 2 + j
        pair_cell = f"Criteria_Pairwise!{ws.cell(row=r, column=c).coordinate}"
        col_sum_cell = f"Criteria_Pairwise!{ws.cell(row=col_sum_row, column=c).coordinate}"
        cn.cell(row=r, column=c, value=f"={pair_cell}/{col_sum_cell}")

# weights = row average
for i in range(n):
    r = 2 + i
    first = cn.cell(row=r, column=2).coordinate
    last = cn.cell(row=r, column=1+n).coordinate
    cn.cell(row=r, column=2+n+1, value=f"=AVERAGE({first}:{last})")  # put weights starting at column n+3
cn.cell(row=1, column=2+n+1, value="Weight")

# lambda_max = SUMPRODUCT(column_sums, weights)
# column_sums in Criteria_Pairwise row col_sum_row, weights in this sheet column n+3 rows 2..n+1
weights_col = 2 + n + 1
col_sum_start = ws.cell(row=col_sum_row, column=2).coordinate
col_sum_end = ws.cell(row=col_sum_row, column=1+n).coordinate
weights_start = cn.cell(row=2, column=weights_col).coordinate
weights_end = cn.cell(row=1+n, column=weights_col).coordinate
cn.cell(row=col_sum_row, column=weights_col, value=f"=SUMPRODUCT(Criteria_Pairwise!{col_sum_start}:{col_sum_end},{weights_start}:{weights_end})")
cn.cell(row=col_sum_row+1, column=weights_col, value="=IF({0}>0,({0}-{1})/({1}-1),0)".format(cn.cell(row=col_sum_row, column=weights_col).coordinate, n))
# Above is placeholder; we'll put CI and CR below differently
cn.cell(row=col_sum_row+1, column=1, value="lambda_max")

# CI and CR using CHOOSE for RI
# put n in a cell
cn.cell(row=col_sum_row+2, column=1, value="n")
cn.cell(row=col_sum_row+2, column=2, value=n)
cn.cell(row=col_sum_row+3, column=1, value="CI")
cn.cell(row=col_sum_row+4, column=1, value="RI")
cn.cell(row=col_sum_row+5, column=1, value="CR")
# lambda_max formula at col_sum_row
lambda_cell = cn.cell(row=col_sum_row, column=weights_col).coordinate
n_cell = cn.cell(row=col_sum_row+2, column=2).coordinate
ci_cell = cn.cell(row=col_sum_row+3, column=2)
ri_cell = cn.cell(row=col_sum_row+4, column=2)
cr_cell = cn.cell(row=col_sum_row+5, column=2)
ci_cell.value = f"=({lambda_cell}-{n_cell})/({n_cell}-1)"
# RI via CHOOSE: index n -> CHOOSE(n,0,0,0.58,0.90,1.12,1.24,1.32,1.41,1.45,1.49)
ri_cell.value = f"=CHOOSE({n_cell},0,0,0.58,0.90,1.12,1.24,1.32,1.41,1.45,1.49)"
cr_cell.value = f"=IF({ri_cell.coordinate}=0,0,{ci_cell.coordinate}/{ri_cell.coordinate})"

# --- Alternatives data ---
alt = wb.create_sheet("Alternatives_Data")
alt_names = ["Alt A","Alt B","Alt C","Alt D"]
alt.cell(row=1, column=1, value="Alternative")
for j, name in enumerate(crit_names, start=2):
    alt.cell(row=1, column=j, value=name)
for i, a in enumerate(alt_names, start=2):
    alt.cell(row=i, column=1, value=a)
    # sample numeric values for each criterion (editable)
    alt.cell(row=i, column=2, value=1000 + i*100)
    alt.cell(row=i, column=3, value=1 + i%3)
    alt.cell(row=i, column=4, value=25 + i*5)

# For each criterion build pairwise alt matrix sheet and compute local priorities
for ci, cname in enumerate(crit_names, start=2):
    sheet_name = f"Alt_Pairwise_{cname}"
    s = wb.create_sheet(sheet_name)
    # headers
    for j, a in enumerate(alt_names, start=2):
        s.cell(row=1, column=j, value=a)
        s.cell(row=j, column=1, value=a)
    # build pairwise formulas referencing Alternatives_Data
    for i_idx, ai in enumerate(alt_names, start=2):
        for j_idx, aj in enumerate(alt_names, start=2):
            # formula = Alternatives_Data!{col}{row_i} / Alternatives_Data!{col}{row_j}
            col_letter = alt.cell(row=1, column=ci).column_letter
            row_i = i_idx
            row_j = j_idx
            # Note: alt data starts at row 2
            s.cell(row=i_idx, column=j_idx, value=f"=Alternatives_Data!{col_letter}{row_i}/Alternatives_Data!{col_letter}{row_j}")
    # column sums
    sum_row = 2 + len(alt_names)
    for j_idx in range(2, 2+len(alt_names)):
        s.cell(row=sum_row, column=j_idx, value=f"=SUM({s.cell(row=2, column=j_idx).coordinate}:{s.cell(row=1+len(alt_names), column=j_idx).coordinate})")
    s.cell(row=sum_row, column=1, value="Column Sum")
    # normalized sheet for this criterion
    ns = wb.create_sheet(f"Alt_Normalized_{cname}")
    for j, a in enumerate(alt_names, start=2):
        ns.cell(row=1, column=j, value=a)
        ns.cell(row=j, column=1, value=a)
    for i_idx in range(2, 2+len(alt_names)):
        for j_idx in range(2, 2+len(alt_names)):
            pair_ref = f"{sheet_name}!{wb[sheet_name].cell(row=i_idx, column=j_idx).coordinate}"
            sum_ref = f"{sheet_name}!{wb[sheet_name].cell(row=sum_row, column=j_idx).coordinate}"
            ns.cell(row=i_idx, column=j_idx, value=f"={pair_ref}/{sum_ref}")
    # local priority = row average
    for i_idx in range(2, 2+len(alt_names)):
        first = ns.cell(row=i_idx, column=2).coordinate
        last = ns.cell(row=i_idx, column=1+len(alt_names)).coordinate
        ns.cell(row=i_idx, column=2+len(alt_names)+1, value=f"=AVERAGE({first}:{last})")
    ns.cell(row=1, column=2+len(alt_names)+1, value="Local Priority")

# Aggregation sheet
agg = wb.create_sheet("Aggregation")
agg.cell(row=1, column=1, value="Alternative")
for i, a in enumerate(alt_names, start=2):
    agg.cell(row=i, column=1, value=a)
# fetch criteria weights from Criteria_Normalized weights column
for j, cname in enumerate(crit_names, start=2):
    agg.cell(row=1, column=j, value=cname)
    # local priorities in Alt_Normalized_cname sheet at column (last+1)
    local_col = 2+len(alt_names)+1
    for i_idx in range(2, 2+len(alt_names)):
        agg.cell(row=i_idx, column=j, value=f"=Alt_Normalized_{cname}!{wb[f'Alt_Normalized_{cname}'].cell(row=i_idx, column=local_col).coordinate}")
# weights
for j, cname in enumerate(crit_names, start=2):
    # weight cell location in Criteria_Normalized: row 2..n+1, weights in column weights_col
    agg.cell(row=2+len(alt_names)+1, column=j, value=f"=Criteria_Normalized!{cn.cell(row=2+j-2, column=weights_col).coordinate}")

# global score as SUMPRODUCT of local priorities row and criteria weights
for i_idx in range(2, 2+len(alt_names)):
    # build range for local priorities across criteria
    local_cells = ",".join([f"{agg.cell(row=i_idx, column=j).coordinate}" for j in range(2, 2+len(crit_names))])
    weight_cells = ",".join([f"{agg.cell(row=2+len(alt_names)+1, column=j).coordinate}" for j in range(2, 2+len(crit_names))])
    # SUMPRODUCT requires ranges; we will place weights in row (2+len(alt_names)+1) so we can reference range
    # create a formula: =SUMPRODUCT(B{row}:D{row}, B{wrow}:D{wrow}) but we need contiguous ranges
    first_local = agg.cell(row=i_idx, column=2).coordinate
    last_local = agg.cell(row=i_idx, column=1+len(crit_names)).coordinate
    first_w = agg.cell(row=2+len(alt_names)+1, column=2).coordinate
    last_w = agg.cell(row=2+len(alt_names)+1, column=1+len(crit_names)).coordinate
    agg.cell(row=i_idx, column=2+len(crit_names)+2, value=f"=SUMPRODUCT({first_local}:{last_local},{first_w}:{last_w})")
agg.cell(row=1, column=2+len(crit_names)+2, value="Global Score")

# Save workbook
out = r"c:\Users\ANAK BAIK\Documents\MK SPK\salinan\spk_kelayakan_kredit\ahp_manual_template.xlsx"
wb.save(out)
print('Saved', out)
